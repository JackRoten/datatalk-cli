"""File format handling with interactive selection for multi-sheet Excel files."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from rich import box
from rich.table import Table

from datatalk.printer import Printer


def detect_excel_sheets(path: str) -> list[str]:
    """Return list of sheet names in an Excel file."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def preview_sheet(path: str, sheet_name: str, max_rows: int = 5) -> pd.DataFrame:
    """Read the first few rows of a specific sheet for preview."""
    return pd.read_excel(path, sheet_name=sheet_name, nrows=max_rows)


def display_sheet_preview(printer: Printer, sheet_name: str, df: pd.DataFrame) -> None:
    """Display a Rich table preview of a sheet's first rows."""
    table = Table(
        title=f"Sheet: {sheet_name}",
        show_header=True,
        header_style="bold magenta",
        box=box.SIMPLE,
    )
    for col in df.columns:
        table.add_column(str(col), style="cyan")
    for _, row in df.iterrows():
        table.add_row(*[str(val) for val in row])
    printer.decorative(table)


def select_excel_sheet(path: str, printer: Printer) -> str | None:
    """Interactively select a sheet from a multi-sheet Excel file.

    Returns the selected sheet name, or None for single-sheet files.
    """
    sheets = detect_excel_sheets(path)

    if len(sheets) <= 1:
        return None

    printer.decorative(
        f"\n[bold yellow]Excel file has {len(sheets)} sheets:[/bold yellow]",
        highlight=False,
    )

    for i, name in enumerate(sheets, 1):
        printer.decorative(f"  [cyan]{i}[/cyan]. {name}", highlight=False)

    printer.decorative("")

    # Show preview of each sheet
    for name in sheets:
        try:
            df = preview_sheet(path, name)
            if not df.empty:
                display_sheet_preview(printer, name, df)
        except Exception:
            printer.decorative(
                f"  [dim]Could not preview sheet '{name}'[/dim]", highlight=False
            )

    # Prompt user to select
    while True:
        try:
            choice = input(f"Select a sheet [1-{len(sheets)}]: ").strip()
        except (EOFError, KeyboardInterrupt):
            return sheets[0]

        if not choice:
            continue

        try:
            idx = int(choice)
            if 1 <= idx <= len(sheets):
                selected = sheets[idx - 1]
                printer.decorative(
                    f"\n[green]Using sheet: {selected}[/green]", highlight=False
                )
                return selected
        except ValueError:
            pass

        printer.decorative(
            f"[red]Please enter a number between 1 and {len(sheets)}[/red]",
            highlight=False,
        )
